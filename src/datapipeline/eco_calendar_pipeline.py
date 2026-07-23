"""从手工下载的月度 Excel 构建经济日历 Excel 和标准 CSV。"""

from __future__ import annotations

from datetime import datetime, time as datetime_time
import os
from pathlib import Path
import re
from typing import Any
from uuid import uuid4

import pandas as pd

from .paths import (
    ECO_CALENDAR_CSV_PATH,
    ECO_CALENDAR_RAW_DIR,
    ECO_CALENDAR_XLSX_PATH,
)
from .pipeline_io import _atomic_write_csv


ECO_SOURCE_COLUMNS = {
    "日期": "date",
    "时间": "time",
    "国家/地区": "region",
    "指标名称": "indicator",
    "重要性": "importance",
    "前值": "prev",
    "预测值": "forecast",
    "今值": "actual",
}
ECO_OUTPUT_COLUMNS = [
    "date",
    "time",
    "datetime",
    "region",
    "indicator",
    "importance",
    "prev",
    "forecast",
    "actual",
    "tf_category",
]


def _normalise_clock(value: Any) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (datetime, pd.Timestamp, datetime_time)):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        total_minutes = int(round(float(value) * 24 * 60)) % (24 * 60)
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    text = str(value).strip()
    match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::\d{2})?", text)
    if match is None:
        return text
    return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}"


def _classify_eco_indicator(indicator: pd.Series) -> pd.Series:
    """将中国重要指标划分为通胀、信用、增长和 PMI 四类。"""
    text = indicator.astype("string")
    category = pd.Series(pd.NA, index=text.index, dtype="string")
    category.loc[
        text.str.contains(r"CPI|PPI", case=False, regex=True, na=False)
    ] = "inflation"
    category.loc[
        text.str.contains(
            r"M0|M1|M2|社会融资|人民币贷款",
            case=False,
            regex=True,
            na=False,
        )
    ] = "credit"
    category.loc[
        text.str.contains("GDP", case=False, regex=False, na=False)
    ] = "growth"
    category.loc[
        text.str.contains("官方制造业PMI", case=False, regex=False, na=False)
    ] = "pmi"
    return category


def _write_eco_calendar_xlsx(frame: pd.DataFrame, path: str | Path) -> Path:
    """生成便于导师查看的 Excel，同时保持日期字段为真实日期类型。"""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.stem}.{uuid4().hex}.tmp.xlsx")
    try:
        with pd.ExcelWriter(
            temporary,
            engine="openpyxl",
            date_format="yyyy-mm-dd",
            datetime_format="yyyy-mm-dd hh:mm:ss",
        ) as writer:
            frame.to_excel(writer, sheet_name="Sheet1", index=False)
            sheet = writer.book["Sheet1"]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            header_fill = PatternFill("solid", fgColor="1F4E78")
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            widths = [12, 9, 20, 12, 42, 12, 14, 14, 14, 15]
            for index, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = width
            for cell in sheet["A"][1:]:
                cell.number_format = "yyyy-mm-dd"
            for cell in sheet["C"][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        if not temporary.is_file() or temporary.stat().st_size == 0:
            raise ValueError(f"Excel 输出为空: {temporary}")
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()
    return target


def build_eco_calendar_files(
    raw_dir: str | Path = ECO_CALENDAR_RAW_DIR,
    xlsx_path: str | Path = ECO_CALENDAR_XLSX_PATH,
    csv_path: str | Path = ECO_CALENDAR_CSV_PATH,
) -> dict[str, Any]:
    """合并月度原始文件，输出 filtered Excel 和 DuckDB 使用的 CSV。"""
    source_dir = Path(raw_dir)
    files = sorted(source_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"经济日历原始目录没有 xlsx: {source_dir}")

    pieces: list[pd.DataFrame] = []
    required = set(ECO_SOURCE_COLUMNS)
    for path in files:
        raw = pd.read_excel(path, sheet_name="经济数据")
        missing = sorted(required - set(raw.columns))
        if missing:
            raise ValueError(f"{path.name}/经济数据 缺少列: {missing}")
        pieces.append(raw.rename(columns=ECO_SOURCE_COLUMNS))

    combined = pd.concat(pieces, ignore_index=True)
    combined["region"] = combined["region"].astype("string").str.strip()
    combined["importance"] = combined["importance"].astype("string").str.strip()
    combined["indicator"] = combined["indicator"].astype("string").str.strip()
    combined["tf_category"] = _classify_eco_indicator(combined["indicator"])
    filtered = combined.loc[
        combined["region"].eq("中国")
        & combined["importance"].eq("重要")
        & combined["tf_category"].notna()
    ].copy()
    filtered["date"] = pd.to_datetime(
        filtered["date"], errors="coerce"
    ).dt.normalize()
    filtered["time"] = filtered["time"].map(_normalise_clock)
    date_text = filtered["date"].dt.strftime("%Y-%m-%d")
    filtered["datetime"] = pd.to_datetime(
        date_text + " " + filtered["time"], errors="coerce"
    )
    invalid = (
        filtered["date"].isna()
        | filtered["time"].eq("")
        | filtered["datetime"].isna()
    )
    if invalid.any():
        sample = filtered.loc[
            invalid, ["date", "time", "indicator"]
        ].head(10)
        raise ValueError(f"经济日历存在无法解析的日期/时间:\n{sample}")
    for column in ("prev", "forecast", "actual"):
        filtered[column] = pd.to_numeric(filtered[column], errors="coerce")

    filtered = (
        filtered.drop_duplicates(
            ["datetime", "region", "indicator"], keep="last"
        )
        .sort_values("datetime", kind="stable")
        .reset_index(drop=True)
    )[ECO_OUTPUT_COLUMNS]
    _atomic_write_csv(filtered, csv_path, ECO_OUTPUT_COLUMNS)
    _write_eco_calendar_xlsx(filtered, xlsx_path)
    return {
        "source_files": len(files),
        "rows": len(filtered),
        "start_date": filtered["date"].min().date(),
        "end_date": filtered["date"].max().date(),
        "xlsx_path": str(Path(xlsx_path).resolve()),
        "csv_path": str(Path(csv_path).resolve()),
    }
